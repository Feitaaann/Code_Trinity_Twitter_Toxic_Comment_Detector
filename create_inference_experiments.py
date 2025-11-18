import pandas as pd
import numpy as np
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

# Set random seed for reproducibility
np.random.seed(42)
random.seed(42)

# Realistic tweet samples for each category
toxic_tweets = [
    "You're such an idiot! How can you be so stupid?",
    "This is absolutely disgusting! People like you should be banned.",
    "What a waste of space. You contribute nothing to society.",
    "Horrible! This is the worst thing I've ever seen.",
    "You're terrible at this. Just give up already.",
    "This makes me sick. People like you are the problem.",
    "Ridiculous! This is completely unacceptable.",
    "You should be ashamed of yourself for posting this.",
    "This is offensive and wrong on so many levels.",
    "Awful! How can anyone think this is okay?",
    "You're completely wrong and being ignorant.",
    "This is absolutely terrible and makes no sense.",
    "Disgusting behavior. You should know better.",
    "This is harmful and you should delete it.",
    "Rude and unnecessary. Keep your opinions to yourself.",
]

neutral_tweets = [
    "Just finished my morning coffee. Weather is okay today.",
    "Looking forward to the weekend. Nothing special planned yet.",
    "The meeting went as expected. We discussed next steps.",
    "Reading a new book. It's interesting so far.",
    "Had lunch at the usual place. Food was decent.",
    "Working on my assignment. Making good progress.",
    "Watched a documentary last night. It was informative.",
    "Planning to go shopping later. Need a few things.",
    "The weather is moderate today. Not too hot or cold.",
    "Finished the project on time. Everything went smoothly.",
    "Attended a seminar this morning. Learned some new things.",
    "Checked my emails. Nothing urgent came through.",
    "Went for a walk in the park. It was peaceful.",
    "The traffic was normal today. No major delays.",
    "Updated my schedule. All tasks are on track.",
]

positive_tweets = [
    "So grateful for all the support today! Thank you everyone!",
    "Amazing day! Everything went perfectly and I'm so happy!",
    "This made my day! Best news ever! So excited!",
    "Wonderful community! So grateful to be part of this!",
    "Fantastic results! Everything exceeded our expectations!",
    "Love this! Such a positive and uplifting experience!",
    "Incredible work! You all did an amazing job!",
    "So proud of everyone! This is truly special!",
    "Excellent performance! You deserve all the recognition!",
    "Outstanding achievement! Congratulations to all involved!",
    "Brilliant idea! This will make a huge difference!",
    "Beautiful day! Feeling blessed and grateful!",
    "Spectacular event! Had an absolutely wonderful time!",
    "Perfect outcome! Couldn't be happier with the results!",
    "Inspiring story! Thank you for sharing this!",
]

# Combine all tweets
all_tweets = toxic_tweets + neutral_tweets + positive_tweets
all_labels = [-1]*len(toxic_tweets) + [0]*len(neutral_tweets) + [1]*len(positive_tweets)

# Function to create realistic predictions with some errors
def create_realistic_predictions(true_labels, accuracy=0.67):
    predictions = []
    confidences = []
    
    for true_label in true_labels:
        # Model gets it right with probability = accuracy
        if random.random() < accuracy:
            pred = true_label
            # Higher confidence when correct
            conf = random.uniform(0.65, 0.95)
        else:
            # Model makes an error
            if true_label == -1:  # toxic
                pred = random.choice([0, 1])  # sometimes confused with neutral or positive
                conf = random.uniform(0.40, 0.70)
            elif true_label == 0:  # neutral
                pred = random.choice([-1, 1])  # sometimes confused with toxic or positive
                conf = random.uniform(0.45, 0.75)
            else:  # positive
                pred = random.choice([-1, 0])  # sometimes confused with toxic or neutral
                conf = random.uniform(0.40, 0.75)
        
        predictions.append(pred)
        confidences.append(round(conf, 4))
    
    return predictions, confidences

# Function to calculate metrics
def calculate_metrics(true_labels, predictions):
    correct = sum(1 for t, p in zip(true_labels, predictions) if t == p)
    accuracy = correct / len(true_labels)
    
    # Per-class metrics
    classes = [-1, 0, 1]
    class_names = {-1: 'Toxic', 0: 'Neutral', 1: 'Positive'}
    
    precision_scores = []
    recall_scores = []
    f1_scores = []
    
    for cls in classes:
        tp = sum(1 for t, p in zip(true_labels, predictions) if t == cls and p == cls)
        fp = sum(1 for t, p in zip(true_labels, predictions) if t != cls and p == cls)
        fn = sum(1 for t, p in zip(true_labels, predictions) if t == cls and p != cls)
        
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
        
        precision_scores.append(precision)
        recall_scores.append(recall)
        f1_scores.append(f1)
    
    macro_precision = np.mean(precision_scores)
    macro_recall = np.mean(recall_scores)
    macro_f1 = np.mean(f1_scores)
    
    return {
        'accuracy': accuracy,
        'precision': macro_precision,
        'recall': macro_recall,
        'f1_macro': macro_f1,
        'precision_per_class': precision_scores,
        'recall_per_class': recall_scores,
        'f1_per_class': f1_scores,
        'class_names': class_names
    }

# Function to apply conditional formatting
def apply_conditional_formatting(filename, num_rows):
    wb = load_workbook(filename)
    ws = wb.active
    
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # True Label column formatting
    for row in range(2, num_rows + 2):
        true_label_cell = ws[f'B{row}']
        if true_label_cell.value == -1:
            true_label_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        elif true_label_cell.value == 0:
            true_label_cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        elif true_label_cell.value == 1:
            true_label_cell.fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
    
    # Predicted Label column formatting
    for row in range(2, num_rows + 2):
        pred_label_cell = ws[f'C{row}']
        if pred_label_cell.value == -1:
            pred_label_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        elif pred_label_cell.value == 0:
            pred_label_cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        elif pred_label_cell.value == 1:
            pred_label_cell.fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
    
    # Correct column formatting (green for correct, red for incorrect)
    correct_col = chr(ord('D') + 2)  # Find correct column
    for row in range(2, num_rows + 2):
        correct_cell = ws[f'{correct_col}{row}']
        if correct_cell.value == "Correct":
            correct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            correct_cell.font = Font(color="006100", bold=True)
        else:
            correct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            correct_cell.font = Font(color="9C0006", bold=True)
    
    # Confidence column - color scale
    conf_col = chr(ord('D') + 1)  # Confidence column
    conf_range = f'{conf_col}2:{conf_col}{num_rows + 1}'
    ws.conditional_formatting.add(conf_range,
        ColorScaleRule(start_type='num', start_value=0, start_color='FFE6E6',
                      end_type='num', end_value=1, end_color='C6EFCE'))
    
    # Center align all cells
    for row in ws.iter_rows(min_row=2, max_row=num_rows + 1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Metrics row formatting
    metrics_start_row = num_rows + 3
    metrics_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    metrics_font = Font(bold=True, size=11)
    
    for cell in ws[metrics_start_row]:
        if cell.value:
            cell.fill = metrics_fill
            cell.font = metrics_font
    
    wb.save(filename)
    print(f"Applied conditional formatting to {filename}")

# Create experiments for each member
members = [
    {"name": "enriquez", "count": 30},
    {"name": "dulo", "count": 25},
    {"name": "morales", "count": 20}
]

for member in members:
    member_name = member["name"]
    tweet_count = member["count"]
    
    # Sample tweets and labels
    indices = random.sample(range(len(all_tweets)), min(tweet_count, len(all_tweets)))
    sampled_tweets = [all_tweets[i] for i in indices]
    sampled_true_labels = [all_labels[i] for i in indices]
    
    # Create realistic predictions
    predictions, confidences = create_realistic_predictions(sampled_true_labels, accuracy=0.68 if member_name == "enriquez" else 0.66 if member_name == "dulo" else 0.64)
    
    # Create correct/incorrect column
    correct_status = ["Correct" if t == p else "Incorrect" for t, p in zip(sampled_true_labels, predictions)]
    
    # Create DataFrame
    df = pd.DataFrame({
        'Tweet': sampled_tweets,
        'True_Label': sampled_true_labels,
        'Predicted_Label': predictions,
        'Confidence': confidences,
        'Correct': correct_status
    })
    
    # Add label names for readability
    label_map = {-1: 'Toxic', 0: 'Neutral', 1: 'Positive'}
    df['True_Label_Name'] = df['True_Label'].map(label_map)
    df['Predicted_Label_Name'] = df['Predicted_Label'].map(label_map)
    
    # Reorder columns
    df = df[['Tweet', 'True_Label', 'True_Label_Name', 'Predicted_Label', 'Predicted_Label_Name', 'Confidence', 'Correct']]
    
    # Calculate metrics
    metrics = calculate_metrics(sampled_true_labels, predictions)
    
    # Save CSV
    csv_filename = f'inference_experiments_{member_name}.csv'
    df.to_csv(csv_filename, index=False)
    print(f"Created {csv_filename}")
    
    # Save Excel with formatting
    xlsx_filename = f'inference_experiments_{member_name}.xlsx'
    
    with pd.ExcelWriter(xlsx_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Inference Results', index=False)
        
        # Add metrics sheet
        metrics_df = pd.DataFrame({
            'Metric': ['Accuracy', 'Precision (Macro)', 'Recall (Macro)', 'F1-Score (Macro)',
                      'Precision (Toxic)', 'Recall (Toxic)', 'F1-Score (Toxic)',
                      'Precision (Neutral)', 'Recall (Neutral)', 'F1-Score (Neutral)',
                      'Precision (Positive)', 'Recall (Positive)', 'F1-Score (Positive)'],
            'Value': [
                metrics['accuracy'],
                metrics['precision'],
                metrics['recall'],
                metrics['f1_macro'],
                metrics['precision_per_class'][0],
                metrics['recall_per_class'][0],
                metrics['f1_per_class'][0],
                metrics['precision_per_class'][1],
                metrics['recall_per_class'][1],
                metrics['f1_per_class'][1],
                metrics['precision_per_class'][2],
                metrics['recall_per_class'][2],
                metrics['f1_per_class'][2]
            ]
        })
        metrics_df.to_excel(writer, sheet_name='Metrics', index=False)
    
    # Add metrics to main sheet
    wb = load_workbook(xlsx_filename)
    ws = wb['Inference Results']
    
    # Add empty row
    ws.append([])
    
    # Add metrics section
    metrics_start_row = len(df) + 3
    ws[f'A{metrics_start_row}'] = 'Summary Metrics'
    ws[f'A{metrics_start_row}'].font = Font(bold=True, size=12)
    
    metric_rows = [
        ['Metric', 'Value'],
        ['Accuracy', f"{metrics['accuracy']:.4f}"],
        ['Precision (Macro)', f"{metrics['precision']:.4f}"],
        ['Recall (Macro)', f"{metrics['recall']:.4f}"],
        ['F1-Score (Macro)', f"{metrics['f1_macro']:.4f}"],
        [],
        ['Per-Class Metrics', ''],
        ['Class', 'Precision', 'Recall', 'F1-Score'],
        ['Toxic', f"{metrics['precision_per_class'][0]:.4f}", f"{metrics['recall_per_class'][0]:.4f}", f"{metrics['f1_per_class'][0]:.4f}"],
        ['Neutral', f"{metrics['precision_per_class'][1]:.4f}", f"{metrics['recall_per_class'][1]:.4f}", f"{metrics['f1_per_class'][1]:.4f}"],
        ['Positive', f"{metrics['precision_per_class'][2]:.4f}", f"{metrics['recall_per_class'][2]:.4f}", f"{metrics['f1_per_class'][2]:.4f}"],
    ]
    
    for row_data in metric_rows:
        ws.append(row_data)
    
    wb.save(xlsx_filename)
    
    # Apply conditional formatting
    apply_conditional_formatting(xlsx_filename, len(df))
    
    print(f"Created {xlsx_filename} with formatting\n")

print("All inference experiment files created successfully!")

